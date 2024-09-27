import os
import openpyxl
import re

INPUT = 'DatabaseSpec.xlsx'
OUTPUT = 'ddl.sql'
DIR = os.path.dirname(__file__)
EXCLUSION_SHEETS = []
TABLE_NAME_CELL = "C4"
COLUMN_INFO_FIRST_ROW = 10

class Column:
  def __init__(self):
    self.name: str = ""
    self.datatype: str = ""
    self.size: str = ""
    self.notnull: str = ""
    self.default: str = ""


def make_sql(sheet):
  # Sheet
  ws = wb[sheet]

  # SQL
  tb_name: str = ws[TABLE_NAME_CELL].value
  
  sql: str = "DROP TABLE IF EXISTS `" + tb_name + "`;\n"
  sql += "CREATE TABLE `" + tb_name + "` (\n"

  pks: list[str] = []
  row = COLUMN_INFO_FIRST_ROW
  while(row < 10000):
    # Terminate cell.
    if ws.cell(row = row, column = 2).value == None:
      break
    
    # Get cells value.
    colname: str = ws.cell(row = row, column = 3).value
    typename: str = ws.cell(row = row, column = 4).value
    size: str = ws.cell(row = row, column = 5).value
    not_null: str = ws.cell(row = row, column = 6).value
    ai: str = ws.cell(row = row, column = 7).value
    pk: str = ws.cell(row = row, column = 8).value
    default_value: str = ws.cell(row = row, column = 9).value

    # Make sql statement.
    sql += "  `" + colname + "` " + typename
    
    if size != None:
       sql += "(" + str(size) + ")"

    if not_null != None:
      sql += " NOT NULL"
    
    if ai != None:
      sql += " AUTO_INCREMENT"
        
    if default_value != None:
       sql += " default " + str(default_value)

    sql +=",\n"
    
    # Set PRIMARY KEY
    if pk != None:
      pks.append(colname)

    row += 1
  
  # PK to sql
  if len(pks) > 0:
    pk_str = ",".join(pks)
    sql += "  PRIMARY KEY (" + pk_str + ")\n"
  else:
    sql = sql[:-2] + "\n"

  # End Table
  sql += ");"

  return sql

# Main
wb = openpyxl.load_workbook(os.path.join(DIR, INPUT))
sheets = wb.sheetnames
tables = []
for sheet in sheets:
  if sheet in EXCLUSION_SHEETS:
    continue
  
  sql = make_sql(sheet)
  tables.append(sql)

# Convert to Strings.
content = "\n\n".join(tables)

# Export DDL file.
with open(os.path.join(DIR, OUTPUT), 'w', encoding='UTF-8') as f:
  f.write(content)
